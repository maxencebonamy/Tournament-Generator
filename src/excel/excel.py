from typing import List, Tuple
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from openpyxl.cell.text import InlineFont 
from openpyxl.cell.rich_text import TextBlock, CellRichText

from utils.colors import get_dark_color, get_light_color


class Excel:

    CELL_WIDTH = 20
    CELL_HEIGHT = 25

    def __init__(self, planning: List[List[int]], teams: List[str], games: List[str]):
        self.__workbook = Workbook()
        self.__planning = planning
        self.__teams = teams
        self.__games = games

        for sheet_name in self.__workbook.get_sheet_names():
            sheet = self.__workbook.get_sheet_by_name(sheet_name)
            self.__workbook.remove_sheet(sheet)
    
    def generate(self) -> None:
        self.__generate_global_planning()
        self.__generate_teams_planning()
        self.__generate_games_planning()
    
    def save(self, path: str) -> None:
        if not path:
            path = "output"
        if not path.endswith(".xlsx"):
            path += ".xlsx"
        self.__workbook.save(f"output/{path}")
    
    def dispose(self) -> None:
        self.__workbook.close()
    
    def __generate_global_planning(self) -> None:
        sheet = self.__workbook.create_sheet("Global Planning")
        sheet.freeze_panes = sheet["B2"]

        for x in range(len(self.__planning)):
            self.__config_cell(sheet.cell(row=1, column=x + 2), value=self.__teams[x], bold=True, color=get_dark_color(x))

        for y in range(len(self.__planning[0])):
            self.__config_cell(sheet.cell(row=y + 2, column=1), value=f"{y + 1}", bold=True)
            for x in range(len(self.__planning)):
                index = self.__planning[x][y]
                self.__config_cell(sheet.cell(row=y + 2, column=x + 2), value=self.__games[index], bg_color=get_light_color(index))

        for x in range(len(self.__planning) + 1):
            sheet.column_dimensions[get_column_letter(x + 1)].width = Excel.CELL_WIDTH
        for y in range(len(self.__planning[0]) + 1):
            sheet.row_dimensions[y + 1].height = Excel.CELL_HEIGHT

    def __generate_teams_planning(self) -> None:
        sheet = self.__workbook.create_sheet("Teams Planning")
        sheet.freeze_panes = sheet["B2"]

        for y in range(len(self.__planning[0])):
            self.__config_cell(sheet.cell(row=y + 2, column=1), value=f"{y + 1}", bold=True)
        
        for x in range(len(self.__planning)):
            self.__config_cell(sheet.cell(row=1, column=2*x + 2), value=self.__teams[x], bold=True, color=get_dark_color(x))
            sheet.merge_cells(start_row=1, start_column=2*x + 2, end_row=1, end_column=2*x + 3)
            for y in range(len(self.__planning[0])):
                opponent = self.__get_opponent(y, x)
                game = self.__planning[x][y]
                rich_text_cell = CellRichText([
                    TextBlock(InlineFont(color="000000"), "VS    "),
                    TextBlock(InlineFont(color=get_dark_color(opponent)), self.__teams[opponent])
                ])
                self.__config_cell(sheet.cell(row=y + 2, column=2*x + 2), value=f"{self.__games[game]}", bg_color=get_light_color(game))
                self.__config_cell(sheet.cell(row=y + 2, column=2*x + 3), value=rich_text_cell, color=get_dark_color(opponent))
        
        for x in range(2*len(self.__planning) + 1):
            sheet.column_dimensions[get_column_letter(x + 1)].width = Excel.CELL_WIDTH
        for y in range(len(self.__planning[0]) + 1):
            sheet.row_dimensions[y + 1].height = Excel.CELL_HEIGHT

    def __generate_games_planning(self) -> None:
        sheet = self.__workbook.create_sheet("Games Planning")
        sheet.freeze_panes = sheet["B2"]

        for y in range(len(self.__planning[0])):
            self.__config_cell(sheet.cell(row=y + 2, column=1), value=f"{y + 1}", bold=True)
            for x in range(len(self.__planning[0])):
                teams = self.__get_teams(y, x)
                if len(teams) >= 2:
                    rich_text_cell = CellRichText([
                        TextBlock(InlineFont(color=get_dark_color(teams[0])), self.__teams[teams[0]]),
                        TextBlock(InlineFont(color="000000"), "    VS    "),
                        TextBlock(InlineFont(color=get_dark_color(teams[1])), self.__teams[teams[1]])
                    ])
                    self.__config_cell(sheet.cell(row=y + 2, column=2*x + 2), value=rich_text_cell, color=get_dark_color(teams[0]))
                    sheet.merge_cells(start_row=y + 2, start_column=2*x + 2, end_row=y + 2, end_column=2*x + 3)
                else:
                    self.__config_cell(sheet.cell(row=y + 2, column=2*x + 2), bg_color="EAEAEA")
                    sheet.merge_cells(start_row=y + 2, start_column=2*x + 2, end_row=y + 2, end_column=2*x + 3)

        for x in range(len(self.__planning[0])):
            self.__config_cell(sheet.cell(row=1, column=2*x + 2), value=self.__games[x], bold=True, bg_color=get_light_color(x))
            sheet.merge_cells(start_row=1, start_column=2*x + 2, end_row=1, end_column=2*x + 3)
        
        for x in range(2*len(self.__planning[0]) + 1):
            sheet.column_dimensions[get_column_letter(x + 1)].width = Excel.CELL_WIDTH
        for y in range(len(self.__planning[0]) + 1):
            sheet.row_dimensions[y + 1].height = Excel.CELL_HEIGHT

    def __config_cell(self, cell: Cell, *, value: str = "", bold: bool = False, italic: bool = False, color: str = "000000", bg_color: str = "FFFFFF"):
        cell.value = value
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(color=color, bold=bold, italic=italic)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def __get_teams(self, slot: int, game: int) -> List[int]:
        return [team for team in range(len(self.__planning)) if self.__planning[team][slot] == game]

    def __get_opponent(self, slot: int, team: int) -> int:
        return [t for  t in range(len(self.__planning)) if t != team and self.__planning[team][slot] == self.__planning[t][slot]][0]
    